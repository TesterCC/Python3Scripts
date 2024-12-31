# -*- coding:utf-8 -*-
__author__ = 'MFC'
__time__ = '18/2/9 15:16'

import os
from openpyxl import Workbook, load_workbook


# excel自定义封装类
class CustomOpenXL:
    def __init__(self, path, read_only=False):
        self.wb = None
        if os.path.exists(path):
            self.path = path
            self.wb = load_workbook(self.path, read_only=read_only)
        else:
            print("%s 文件不存在" % path)
            exit(0)

    # 获取excel的行数
    # 如果指定的工作簿存在，则返回其数据行数，否则返回None
    def get_cell_row(self, sheet):
        if self.wb:
            # 先通过sheet获取工作簿
            sh = self.wb.get_sheet_by_name(sheet)
            if sh:
                return sh.max_row

        return None

    # 获取excel的列数
    # 如果指定的工作簿存在，则返回其数据列数，否则返回None
    def get_cell_col(self, sheet):
        if self.wb:
            # 先通过sheet获取工作簿
            sh = self.wb.get_sheet_by_name(sheet)
            if sh:
                return sh.max_column

        return None

    # 获取工作簿名称列表
    def get_sheets_name(self):
        if self.wb:
            return self.wb.get_sheet_names()

        return None

    # 通过索引获取工作簿名
    # 索引从0开始
    def get_sheet_name_by_index(self, index):
        if self.wb:
            sheets = self.wb.get_sheet_names()
            sheet_len = len(sheets)
            if index >= 0 and index < sheet_len:
                return sheets[index]

        return None

    # 创建工作簿
    def create_sheet(self, name, index=0):
        res = False
        if self.wb:
            self.wb.create_sheet(title=name, index=index)
            res = True

        return res

    # 修改工作簿名
    def set_sheet_name(self, sheet_name, name):
        res = False
        if self.wb:
            self.wb[sheet_name].title = name
            res = True

        return res

    # 获取单元格值
    def get_cell_value(self, sheet, row, col):
        value = None
        if self.wb:
            value = self.wb[sheet].cell(row=row, column=col).value

        return value

    # 设置单元格值
    def set_cell_value(self, sheet, row, col, value):
        res = False
        if self.wb:
            self.wb[sheet].cell(row=row, column=col).value = value
            res = True

        return res

    # 保存
    def save(self, path=""):
        if path != "":
            self.path = path

        if self.wb:
            self.wb.save(self.path)
