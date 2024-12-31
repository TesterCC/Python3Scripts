# coding=utf-8
"""
DATE:   2021/11/18
AUTHOR: TesterCC
"""
import os

from openpyxl import Workbook
from openpyxl import load_workbook


# 比较快捷的还是用csv方式处理，还是name_tuple csv处理方式比较方便

file_path = r"E:\github_repo\Python3Scripts\dev_demo\modify_file_name\150个武器分类表_L4.xlsx"

file_name = "150个武器分类表_L4.xlsx"

dir_path = r"C:\Users\Yilan\Desktop\search\exp"

def excel_read(filename):
    """
    读取已存在的excel文件
    读取excel指定列内容，写入txt
    """
    print("读取已存在的excel文件")
    wb = load_workbook(filename)

    # 获取所有sheet名, 返回的是list类型
    print("[DEBUG]", wb.sheetnames)
    sheets = wb.sheetnames   # list
    # print(type(sheets))

    # 遍历sheets，并读取其单元格内容打印输出
    for sh in sheets:
        print("读取工作簿名称： ", sh)

    # 获取要读取的sheet
    ws = wb['cve']

    # print(type(ws.rows))
    # for row in ws.rows:
    #     print(row[2].value)


    origin_cve_no = [row[2].value for row in ws.rows if row[2].value and row[2].value != 'CVE编号']
    # print(f"长度: {len(origin_cve_no)} \n{origin_cve_no}")

    with open("150_cve_no.txt", "w+") as f:
        for i in origin_cve_no:
            f.write(i + "\n")

    return origin_cve_no

def read_dir_file_name(dir_path):
    """
    获取文件夹下文件的名称
    """
    # for root, dirs, files in os.walk(dir_path):
    #     print(files) # 当前路径下所有非目录子文件

        # print(root)  # 当前目录路径
        # print(dirs)  # 当前路径下所有子目录
        # print(files) # 当前路径下所有非目录子文件
    file_name_list = [i.lower() for root, dirs, files in os.walk(dir_path) for i in files]
    # print(file_name_list)
    return file_name_list

def match_file():
    # 筛选出匹配的cve_no，这些cve文件需要改名
    origin_cve_no = excel_read(file_name)
    cve_no_150 = [ i.lower() for i in origin_cve_no]

    file_name_list = read_dir_file_name(dir_path)

    print(cve_no_150)
    print(file_name_list)

    for cve_no in cve_no_150:
        for fn in file_name_list:
            if cve_no in fn:
                print(cve_no)
                # todo rename file  但这里这样处理逻辑不佳

    match_list = [ cve_no.upper() for cve_no in cve_no_150 for fn in file_name_list if cve_no in fn]
    # print(match_list)
    # print(len(match_list))

    return match_list


if __name__ == '__main__':
    # origin_cve_no = excel_read(file_name)
    # cve_no_150 = [ i.lower() for i in origin_cve_no]
    #
    # print(cve_no_150)
    match_list = match_file()
    print(match_list)


