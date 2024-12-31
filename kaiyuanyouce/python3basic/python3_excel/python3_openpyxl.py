# -*- coding:utf-8 -*-
__author__ = 'MFC'
__time__ = '18/1/17 22:31'

"""
读写Excel - openpyxl库
http://openpyxl.readthedocs.io/en/default/

openpyxl是一个第三方的pythonexcel读写库，支持Excel2010 xlsx/xlsm/xltx/xltm文件格式。

https://mp.weixin.qq.com/s?__biz=MzI0NDQ5NzYxNg==&mid=2247484065&idx=1&sn=35cb15d02a6d503714f9e220397eb438&scene=19#wechat_redirect
"""


from openpyxl import Workbook
from openpyxl import load_workbook


def excel_write():
    """
    写excel简单示例
    """
    print("写excel简单示例")

    # 创建一个excel工作区
    wb = Workbook()

    # 获取sheet名称
    print(wb.get_sheet_names())

    # 激活当前工作簿 worksheet
    ws = wb.active

    # 往单元格A1写入数据, 其他单元格写入类似
    ws['A1'] = "测试开发"

    # 在第2行写下一横行数据，列表元素对应每一个单元格 横行纵列
    ws.append([1, 2, 3])

    # 写入时间类型到excel, python会自动将类型转换成excel的日期时间类型
    from datetime import datetime
    ws['A2'] = datetime.now()   # A2的数字1被替换为时间

    # 保存为excel文件
    wb.save("Excel写入sample.xlsx")


def excel_read(filename):
    """
    读取已存在的excel文件
    """
    print("读取已存在的excel文件")
    wb = load_workbook(filename)

    # 获取所有sheet名, 返回的是list类型
    sheets = wb.get_sheet_names()
    print(type(sheets))

    # 遍历sheets，并读取其单元格内容打印输出
    for sh in sheets:
        print("读取工作簿名称： ", sh)

    # 获取要读取的sheet
    ws = wb.get_sheet_by_name(sheets[0])

    # 读取Sheet A1 和 A2, B2, C2单元格内容
    # 读取A1单元格的值
    A1 = ws['A1'].value
    print("A1单元格的值： ", A1)

    # 读取A2, B2, C2
    for index in ('A2', 'B2', 'C2'):
        print(index, "单元格的值： ", ws[index].value)

    # 读取空值的单元格, openpyxl对于空值的单元格，返回None
    F1 = ws['F1'].value
    print("F1单元格的值(空单元格)： ", F1)


def douban_spider_save_excel(filename):
    """
    使用urllib从网络爬取数据，写入excel进行示例演示，从豆瓣网爬取部分书籍数据，写入excel
    通过豆瓣网搜索API，搜索python关键词，采集书籍数据
    本示例只采集第一页的数据
    """
    import urllib.request

    url = "https://api.douban.com/v2/book/search?q=python"
    response = urllib.request.urlopen(url)

    # 将bytes数据流解码成string
    ebook_str = response.read().decode()

    # 将string转换成dict
    ebook_dict = eval(ebook_str)

    # 构建一个Workbook对象
    wb = Workbook()

    # 激活第一个sheet
    ws = wb.active

    # 写入表头
    ws.append(["书名", "作者", "描述", "出版社", "价格"])

    # 写入书信息
    print("开始写入书信息")
    for book in ebook_dict["books"]:
        ws.append([book["title"],
                   ",".join(book["author"]),
                   book["summary"],
                   book["publisher"],
                   book["price"]])

    # 保存
    wb.save(filename)
    print("当前页书籍信息已写入Excel")


if __name__ == '__main__':
    # excel_write()
    # excel_read("Excel写入sample.xlsx")
    douban_spider_save_excel("ebook.xlsx")




