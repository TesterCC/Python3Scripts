# -*- coding:utf-8 -*-
__author__ = 'MFC'
__time__ = '18/1/31 22:06'

"""
扩展

请根据上面任务练习，进一步深入学习openpyxl库，掌握以下：

合并单元格读写、操作
单元格背景设置
单元格字体、颜色设置
图表生成


查阅openpyxl官方文档去实现以下功能:
一次性读取多个单元格的值
设置指定单元格的字体
设置指定单元格的样式（例如背景色等）

（字体、填充、边框、位置和保护）实例一旦被创建实例的属性就不可更改，只能重新创建实例。
若要改变样式，必须新建样式实例
"""

from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import numbers
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 给默认的工作簿修改名称
ws.title = "测试样式设置表"

a1 = ws['A1']
d4 = ws['D4']

ft = Font(color=colors.RED, size=32, bold=True, underline="single")
a1.font = ft
d4.font = ft

# 对第一行,A-F列写入数据
for col in ("A", "B", "C", "D", "E", "F"):
    ws["%s1" % col] = "测试开发"  # A1 B1 ...

for col in ("A", "B", "C", "D", "E", "F"):
    ws["%s2" % col] = "测试开发2"  # A1 B1 ...

for col in ("A", "B", "C", "D", "E", "F"):
    ws["%s3" % col] = "测试开发3"  # A1 B1 ...

for col in ("A", "B", "C", "D", "E", "F"):
    ws["%s4" % col] = "测试开发4"  # A1 B1 ...


# 若要改变样式，必须新建样式实例
# 创建字体实例（红色、斜体，其他属性默认）
a1.font = Font(color=colors.BLUE, size=22, italic=True)

# 保存excel文档到硬盘
wb.save('openpyxl_style_demo.xlsx')

# 数据格式属性number_format的值是字符串类型，不为对象，直接赋值即可。
# method1
# ws.cell["D1"].number_format = numbers.FORMAT_GENERAL

# method2
ws.cell(row=2, column=4).number_format = numbers.FORMAT_DATE_XLSX15

# method3 直接使用字符串
ws.cell(row=3, column=4).number_format = 'd-mmm-yy'











