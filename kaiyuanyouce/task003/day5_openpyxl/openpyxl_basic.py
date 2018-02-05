# -*- coding:utf-8 -*-
__author__ = 'MFC'
__time__ = '18/2/5 03:47'


"""
excel如何应用到我们的测试场景中来，通常的应用场景有：

1.用于测试数据的管理维护

2.用于自动化测试（含UI级、接口级等等）用例管理

3.用于测试报告生成

https://mp.weixin.qq.com/s?__biz=MzI0NDQ5NzYxNg==&mid=2247484118&idx=1&sn=66fe71986e00e83e31cccc2ff91fbeb3&scene=19#wechat_redirect

"""

from openpyxl import Workbook, load_workbook


def write_excel():
    print("python openpyxl基本实例演示：")

    # 创建excel文档
    wb = Workbook()
    ws = wb.active

    # 给默认的工作簿修改名称
    ws.title = "默认创建的工作簿"

    # 对第一行,A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws["%s1" % col] = "测试开发"   # A1 B1 ...

    # 对第二行,A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws["%s2" % col] = "公众号: FullStackPentest"   # A2 B2 ...

    # 创建一个工作簿
    ws1 = wb.create_sheet("新创建的工作簿1")

    # 对第一行,A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws1["%s1" % col] = "测试开发1"

    # 对第二行,A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws1["%s2" % col] = "大数据测试"

        # 创建一个工作簿
    ws1 = wb.create_sheet("新创建的工作簿2")

    # 对第一行,A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws1["%s1" % col] = "测试开发2"

    # 对第二行,A-F列写入数据
    for col in ("A", "B", "C", "D", "E", "F"):
        ws1["%s2" % col] = "Python3"

    # 保存excel文档到硬盘
    wb.save('openpyxl_demo.xlsx')


def read_excel():
    # 读取openpyxl_demo.xlsx文档中的内容
    # 只读模式打开
    r_wb = load_workbook(filename='openpyxl_demo.xlsx',
                         read_only=True)

    # 获取所有工作簿名称
    sheets = r_wb.get_sheet_names()

    # 遍历各个工作簿中的内容
    # 即上述写入到第一、二行A-F列的数据
    for sheet in sheets:
        ws = r_wb[sheet]
        print("---" * 20)
        print(">>>读取", sheet)

        # 遍历第一、二行A-F列的数据
        for row in (1, 2):
            for col in ("A", "B", "C", "D", "E", "F"):
                print(ws["%s%d" % (col, row)].value, end='      ')
            print(end='\n')

            # 一次性读取多个单元格的数据
    print(end='\n\n')
    print(">>> 一次性读取多个单元格数据")

    for sheet in sheets:
        ws = r_wb[sheet]
        print("---" * 20)
        print(">>>读取", sheet)
        cells_range = ws["A1": "F2"]
        # 遍历下已读取的单元格的内容
        for cells in cells_range:

            for cell in cells:
                print(cell.value, end='  ')

        print(end='\n')


if __name__ == "__main__":
    write_excel()
    read_excel()
