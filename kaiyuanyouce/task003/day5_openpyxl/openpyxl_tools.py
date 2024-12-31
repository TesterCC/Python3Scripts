#!/usr/bin/env python
# -*- coding:utf-8 -*-
__author__ = 'MFC'
__time__ = '18/2/5 03:58'

"""
https://mp.weixin.qq.com/s?__biz=MzI0NDQ5NzYxNg==&mid=2247484119&idx=1&sn=e3e8f9f7d71146db8f16786fdc9e273f&scene=19#wechat_redirect
基于openpyxl封装一个excel解析类.
请注意，不采用Python的任何高级特性，就简简单单的一个类，实现excel的一些基本操作，并演示如何使用该类。
"""

import os
from openpyxl import Workbook, load_workbook


class ToolOpenXL:
    """
    excel自定义封装类
    """
    def __init__(self, path, read_only=False):
        self.wb = None
        if os.path.exists(path):
            self.path = path
            self.wb = load_workbook(self.path, read_only=read_only)
        else:
            print("%s文件不存在" % path)
            exit(0)

    def get_cell_row(self, sheet):
        """
        获取excel的行数
        如果指定的工作簿存在，则返回其数据行数，否则返回None
        """
        if self.wb:
            # 先通过sheet获取工作簿
            sh = self.wb.get_sheet_by_name(sheet)
            if sh:
                return sh.max_row
        return None

    def get_cell_col(self, sheet):
        """
        获取excel的列数
        如果指定的工作簿存在，则返回其数据列数，否则返回None
        """
        if self.wb:
            # 先通过sheet获取工作簿
            sh = self.wb.get_sheet_by_name(sheet)
            if sh:
                return sh.max_column
        return None

    def get_sheets_name(self):
        """
        获取工作簿名称列表
        """
        if self.wb:
            return self.wb.get_sheet_names()
        return None

    def get_sheet_name_by_index(self, index):
        """
        通过索引获取工作簿名
        索引从0开始
        :param index:
        """
        if self.wb:
            sheets = self.wb.get_sheet_names()
            sheet_len = len(sheets)

            if 0 <= index < sheet_len:
                return sheets[index]
        return None

    def create_sheet(self, name, index=0):
        """
        创建工作簿
        :param name:
        :param index:
        """
        res = False
        if self.wb:
            self.wb.create_sheet(title=name, index=index)
            res = True

        return res

    def set_sheet_name(self, sheet_name, name):
        """
        修改工作簿名
        :param sheet_name:
        :param name:
        """
        res = False
        if self.wb:
            self.wb[sheet_name].title = name
            res = True
        return res

    def get_cell_value(self, sheet, row, col):
        """
        获取单元格值
        :param sheet:
        :param row:
        :param col:
        """
        value = None
        if self.wb:
            value = self.wb[sheet].cell(row=row, column=col).value
        return value

    def set_cell_value(self, sheet, row, col, value):
        """
        设置单元格值
        :param sheet:
        :param row:
        :param col:
        :param value:
        """
        res = False
        if self.wb:
            self.wb[sheet].cell(row=row, column=col).value = value
            res = True
        return res

    def save(self, path=""):
        """
        保存
        :param path:
        """
        if path != "":
            self.path = path

        if self.wb:
            self.wb.save(self.path)


if __name__ == '__main__':
    print("python openpyxl基本实例封装后演示:")
    print("---" * 30)

    xl = ToolOpenXL("openpyxl_demo.xlsx")

    # 获取所有工作簿名
    sheets = xl.get_sheets_name()
    print(">>>获取工作簿列表：", end='\n')
    print(sheets)
    print("---" * 30)

    # 通过索引获取工作簿名
    print(">>>通过索引获取工作簿名：")
    for index in range(0, len(sheets)):
        print(xl.get_sheet_name_by_index(index), end='  ')
    print(end="\n")
    print("---" * 30)

    # 获取各工作簿数据行列数
    for sheet in sheets:
        nrows = xl.get_cell_row(sheet)
        ncols = xl.get_cell_col(sheet)
        print("工作簿[%s]的数据行列数为(%d, %d)" % (sheet, nrows, ncols))
    print("---" * 30)

    # 获取各工作簿中的数据
    print(">>>获取工作簿中数据:")
    for sheet in sheets:
        nrows = xl.get_cell_row(sheet)
        ncols = xl.get_cell_col(sheet)
        print("***" * 20, end='\n')
        print("工作簿[%s]数据如下： " % sheet)
        for row in range(1, nrows+1):
            for col in range(1, ncols+1):
                value = xl.get_cell_value(sheet, row, col)
                print("[%d, %d]->%s" % (row, col, value), end='\t')

    print("---" * 30)

    # 修改各工作簿第一行的数据为：PenTest
    print(">>>设置工作簿中数据")
    for sheet in sheets:
        ncols = xl.get_cell_col(sheet)
        for col in range(1, ncols+1):
            xl.set_cell_value(sheet, row=1, col=col, value="Pentest")

    # 保存
    xl.save()

    print("---" * 30)
    print(">>>获取某一行数据：")
    xl.get_one_row_value(2)













